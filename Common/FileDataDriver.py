
# 用来读取和写入excel
from TaokeEms.config import *
import openpyxl


class FileDataDriver:
    @staticmethod  # 静态方法装饰器：直接通过类名调用，不要self
    def readExcel(file_path=CASEDATAURL, sheet_name=SHEETNAME):
        # 打开现有的Excel文件或创建新的文件
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            print('excel正确打开了')
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            print('没有正确打开excel')

        # 选择或创建指定的工作表
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)

        # 获取列名, 获取所有的key  ，并且遍历变成一个列表
        headers = [cell.value for cell in worksheet[2]]
        # print("所有的header（key）", headers)

        # 将数据存储到列表当中
        data = []
        # 把小的数据从第三行开始
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            # print("当前row", row)
            data.append(dict(zip(headers, row)))
        # print("当前的所有数据：", data)
        workbook.close()
        # print(data)
        return data


    @staticmethod
    def writeDataToExcel(file_path=CASEDATAURL, sheet_name=SHEETNAME, row=None, column=None, value=None):
        # 打开现有的Excel文件或创建新的文件
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # 选择或创建指定的工作表
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)

        # 写入数据到指定行和列
        worksheet.cell(row=row, column=column).value = value

        # 保存修改后的文件 -- 文件不能是打开的状态，不然写入会失败
        workbook.save(file_path)


if __name__ == "__main__":
    print(FileDataDriver.readExcel())



